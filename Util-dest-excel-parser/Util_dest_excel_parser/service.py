
import pandas as pd
import numpy as np
import math
import ast
from Util_dest_excel_parser.dto.modifyRequestDto import *
from openpyxl import load_workbook




# ------------ 필드 명 ------------

COL_NAME = "name"
COL_TAGNAME = "tagNames"
COL_WEIGHT = "weight"





def readFile(file) :
    return getTagsFromExcel(file)

def getTagsFromExcel(file):
    """
    excel 파일로부터 태그 정보만 추출해 Dictionary로 변환합니다.
    """
    df = pd.read_excel(file)
    obj = df.to_dict(orient='records')
    for item in obj:
        item[COL_TAGNAME] = ast.literal_eval(item[COL_TAGNAME])
        item[COL_WEIGHT] = ast.literal_eval(item[COL_WEIGHT])
    return replace_nan_with_none(obj)

def replace_nan_with_none(data):
    """
    재귀적으로 딕셔너리나 리스트 내부의 NaN 값을 None으로 변환하는 함수.
    """
    if isinstance(data, dict):
        return {k: replace_nan_with_none(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [replace_nan_with_none(item) for item in data]
    elif isinstance(data, float) and math.isnan(data):
        return None
    else:
        return data




    

def modifyFile(file, requestArr) :
    
    # ======= 파일 로드 =======
    
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
    load_wb = load_workbook(file, data_only=True)
    load_ws = load_wb['Sheet1']

    # ======= 필드 존재 확인 =======
    
    col_name = findColomnIdx(load_ws, COL_NAME)
    col_tag = findColomnIdx(load_ws, COL_TAGNAME)
    col_weight = findColomnIdx(load_ws, COL_WEIGHT)
    
    if col_name == -1 or col_tag == -1 or col_weight == -1 :
        return None

    # ======= 행 사이즈 확인 =======

    row_first, row_end = findStartEndRow(load_ws)
    row_start = row_first + 1

    if (row_first == -1 or row_end == -1) :
        return None

    # ======= 실제 처리 =======

    for req in requestArr :
        rowNum = req.row + row_first
        if rowNum < row_start or row_end <= rowNum :
            return None

        if (req.task == TaskType.REMOVE) :
            taglist = ast.literal_eval(load_ws.cell(rowNum, col_tag).value)
            weightlist = ast.literal_eval(load_ws.cell(rowNum, col_weight).value)
            
            idx = -1
            try :
                idx = taglist.index(req.tag)
            except :
                continue
            
            del taglist[idx]
            del weightlist[idx]

            load_ws.cell(rowNum, col_tag).value = str(taglist)
            load_ws.cell(rowNum, col_weight).value = str(weightlist)
    
    return load_wb

def isEmptyData(cellData) :
    return (None == cellData or not cellData.strip())

# 시작하는 행(포함)과 끝나는 행(미포함) 튜플을 반환합니다.
def findStartEndRow(sheet) :
    MAX_SKIP_ROWS = 100000
    startr = 1
    while isEmptyData(sheet.cell(startr, 1).value) :
        if (MAX_SKIP_ROWS < startr) :
            print('시작하는 행을 찾을 수 없습니다!')
            return (-1, -1)
        startr = startr + 1
    endr = startr
    while not isEmptyData(sheet.cell(endr, 1).value) :
        endr = endr + 1
    return (startr, endr)
    
def findColomnIdx(sheet, colName) :
    c = 1
    while True :
        cellValue = sheet.cell(1, c).value
        if isEmptyData(cellValue) :
            return -1
        if cellValue == colName :
            return c
        c = c + 1
